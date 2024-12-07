from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from time import sleep
import openpyxl
import re
import numpy as np
import bid
import answer

def wait_url(driver : webdriver.Chrome, url : str):
    print(url)
    while True:
        cur_url = driver.current_url
        if cur_url == url:
            break
        sleep(0.1)
        
def find_element(driver : webdriver.Chrome, whichBy, unique : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return element

def find_elements(driver : webdriver.Chrome, whichBy, unique : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(whichBy, unique)
            break
        except:
            pass
        sleep(0.1)
    return elements

def get_url(string):
    skillset = string.split(",")
    length = len(skillset)
    for i in range(length):
        skillset[i] = skillset[i].strip().replace(" ", "-").replace(".", "-").lower()
        if skillset[i] == "web-design":
            skillset[i] = "web-design-2"
        if skillset[i] == "web-development":
            skillset[i] = "web-development-1"
        if skillset[i] == "e-commerce":
            skillset[i] = "e-commerce-1"
    # url = "https://www.workana.com/jobs?category=it-programming&client_history=1&publication=1d&skills="
    # url = "https://www.workana.com/jobs?category=it-programming&language=xx&publication=1w"
    url="https://www.workana.com/jobs?client_history=1&language=xx"
    # url = "https://www.workana.com/jobs?category=it-programming&language=xx"
    for i in range(length):
        if i == 0:
            url = url + skillset[i]
        else:
            url = url + "%2C" + skillset[i]
    print(url)
    return url

def get_budget(string):
    string = string.replace(",", "")
    numbers = re.findall(r"\d+", string)
    numbers = np.array(numbers, dtype=int)
    get_value = np.max(numbers)
    return int(get_value)

def get_skill(skills):
    if "Python" in skills:
        proper_skill = "Python"
    elif "React.js" in skills or "JavaScript" in skills:
        proper_skill = "React.js"
    elif "Android" in skills:
        proper_skill = "Android"
    elif "PHP" in skills or "Laravel" in skills:
        proper_skill = "php"
    else:
        proper_skill = "Perfect"
    return proper_skill
        
def get_bid(skill, country):
    if skill == "Python":
        if country == "Brazil" or country == "Portugal":
            proper_bid = bid.python_brazil()
        else:
            proper_bid = bid.python_spanish()
    elif skill == "React.js":
        if country == "Brazil" or country == "Portugal":
            proper_bid = bid.react_brazil()
        else:
            proper_bid = bid.react_spanish()
    elif skill == "php":
        if country == "Brazil" or country == "Portugal":
            proper_bid = bid.php_brazil()
        else:
            proper_bid = bid.php_spanish()
    elif skill == "Android":
        if country == "Brazil" or country == "Portugal":
            proper_bid = bid.android_brazil()
        else:
            proper_bid = bid.android_spanish()
    else:
        if country == "Brazil" or country == "Portugal":
            proper_bid = bid.perfect_brazil()
        else:
            proper_bid = bid.perfect_spanish()
    return proper_bid
    
def get_answer(question, country):
    if "availability_to_start" in question:
        if country == "Brazil" or country == "Portugal":
            proper_answer = answer.start_brazil()
        else:
            proper_answer = answer.start_spanish()
    elif "delivery_time_needed" in question:
        if country == "Brazil" or country == "Portugal":
            proper_answer = answer.duration_brazil()
        else:
            proper_answer = answer.duration_spanish()
    elif "experience_on_this_type_of_projects" in question:
        if country == "Brazil" or country == "Portugal":
            proper_answer = answer.experience_brazil()
        else:
            proper_answer = answer.experience_spanish()
    elif "data_needed_to_start" in question:
        if country == "Brazil" or country == "Portugal":
            proper_answer = answer.require_brazil()
        else:
            proper_answer = answer.require_spanish()
    elif "chosen_one_reason" in question:
        if country == "Brazil" or country == "Portugal":
            proper_answer = answer.reason_brazil()
        else:
            proper_answer = answer.reason_spanish()
    else:
        if country == "Brazil" or country == "Portugal":
            proper_answer = answer.extra_brazil()
        else:
            proper_answer = answer.extra_spanish()
    return proper_answer

driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.workana.com/en/login")
wait_url(driver, "https://www.workana.com/en/login")

# remote_debugging_address = "127.0.0.1:9224"
# chrome_options = webdriver.ChromeOptions()
# chrome_options.add_experimental_option("debuggerAddress", remote_debugging_address)
# driver = webdriver.Chrome(options=chrome_options)

while True:
    response = input('Accepted Cookies? (y/n): ')
    if response =='y':
        break
    sleep(1)

bid_num = int(input("How many bids do you want? : "))
minimum_budget = int(input("Your minimum budget(USD)? : "))
want_skill = input("Do you want your own skillset? (y/p/n): ")
if want_skill == 'p':
    url = "https://www.workana.com/jobs?category=it-programming&client_history=1&publication=1d&skills=python%2Creact-js%2Cphp%2Candroid%2Chtml"
elif want_skill == 'y':
    skillset = input("Input your skillset: ")
    url = get_url(skillset)
else:
    # url = "https://www.workana.com/jobs?category=it-programming&client_history=1&publication=1d"
    # url="https://www.workana.com/jobs?category=it-programming&client_history=1&language=xx&publication=3d"
    # url="https://www.workana.com/jobs?client_history=1&language=xx"
    url = "https://www.workana.com/jobs?category=it-programming&language=xx&subcategory=web-development%2Cmobile-development%2Cdata-science-1%2Cdesktop-apps%2Cartificial-intelligence-1%2Cothers-5"
excel = openpyxl.load_workbook("workana.xlsx")
sheet = excel.active
email = sheet.cell(row = 1, column = 1).value
password = sheet.cell(row = 1, column = 2).value
email = sheet.cell(row = 1, column = 1).value
password = sheet.cell(row = 1, column = 2).value

find_element(driver, By.NAME, "email").send_keys(email)
find_element(driver, By.NAME, "password").send_keys(password)
find_element(driver, By.XPATH, "//button[@type='submit']").click()
print("Successfully Logined!")

i = 1
contract_type = 1
while i <= bid_num:
    driver.get(url)
    sleep(1)
    find_element(driver, By.CSS_SELECTOR, 'label[for="language-0"]').click()
    if contract_type == 0:
        find_element(driver, By.XPATH, '//*[@id="block-filters"]/div/div/div[5]/div/div[1]/div/div[3]/label').click()
    else:
        find_element(driver, By.XPATH, '//*[@id="block-filters"]/div/div/div[5]/div/div[1]/div/div[2]/label').click()
        sleep(1)
        find_element(driver, By.XPATH, '//*[@id="block-filters"]/div/div/div[5]/div/section/div/span[2]/div/span[1]/div/div/input').send_keys(minimum_budget)
    sleep(2)
    driver.execute_script("window.scrollBy(0, -500)")
    sleep(1)
    projects = find_elements(driver, By.CLASS_NAME, 'project-item')
    status = False
    for project in projects:
        if not project.find_element(By.CLASS_NAME, 'btn-group').text == "Place a bid":
            print("Already placed a bid.")
            continue
        budget_range = project.find_element(By.CLASS_NAME, 'budget').find_element(By.CLASS_NAME, 'values').text
        try:
            budget = get_budget(budget_range)
        except:
            print("Unknown value.")
            continue
        if budget >= 50 and budget < minimum_budget:
            print("This project's budget({}) is too low.".format(budget))
            continue
        skills = []
        try:
            skill_set = project.find_element(By.CLASS_NAME, 'skills').find_elements(By.TAG_NAME, 'h3')
            for skill in skill_set:
                skills.append(skill.text)
        except:
            pass
        if len(skills) == 0:
            print("No required skill.")
            continue
        proper_skill = get_skill(skills)
        
        # Exceptional
        # if proper_skill == "Android":
        #     print("Sorry, but an exceptional skill.")
        #     continue
        
        country = find_element(project, By.CLASS_NAME, 'country-name').find_element(By.TAG_NAME, 'a').text
        print("Relevant skills: ", skills)
        print("Budget for apply: ", budget)
        print("Client's country: ", country)
        print("Skill for apply: ", proper_skill)
        
        while True:
            try:
                find_element(project, By.CLASS_NAME, 'btn-group').click()
                break
            except:
                sleep(1)
        
        while True:
            try:
                all_sheets = driver.find_element(By.ID, 'bid-add').find_element(By.CLASS_NAME, 'row').find_element(By.CLASS_NAME, 'col-md-9').find_elements(By.CSS_SELECTOR, 'div[class="box-common project-bid"]')
                break
            except:
                sleep(1)
                
        try:
            input_sheets = all_sheets[0].find_elements(By.TAG_NAME, 'span')
            for input_sheet in input_sheets:
                question = input_sheet.find_element(By.TAG_NAME, 'label').get_attribute("for")
                answer_ = get_answer(question, country)
                input_sheet.find_element(By.TAG_NAME, 'input').send_keys(answer_)
        except:
            pass
        skill_sheets = all_sheets[1].find_element(By.CSS_SELECTOR, 'div[class="project-bid-skills"]').find_element(By.CSS_SELECTOR, 'div[class="form-group untouched pristine required"]')
        try:
            recommanded_skills = skill_sheets.find_element(By.CSS_SELECTOR, 'div[class="display-selector"]').find_elements(By.TAG_NAME, 'label')
            for recommanded_skill in recommanded_skills:
                recommanded_skill.find_element(By.TAG_NAME, 'span').click()
        except:
            pass
        try:
            other_skills = skill_sheets.find_element(By.CSS_SELECTOR, 'div[class="search multi-select"]')
            other_skills.find_element(By.TAG_NAME, 'input').click()
            sleep(0.5)
            other_skills.find_element(By.CSS_SELECTOR, 'div[class="multi-select-results"]').find_elements(By.TAG_NAME, 'li')[0].find_element(By.TAG_NAME, 'span').click()
        except:
            pass
        try:
            tasks_sheets = all_sheets[4].find_element(By.CLASS_NAME, 'bid-tasks').find_elements(By.CSS_SELECTOR, 'div[class="bid-tasks-edit"]')
            for tasks_sheet in tasks_sheets:
                select_item = tasks_sheet.find_element(By.TAG_NAME, 'select')
                select_item = Select(select_item)
                select_item.select_by_value("1")
                # print("Included selected.")
        except:
            pass
        find_element(driver, By.ID, 'Amount').send_keys(budget)
        my_bid = get_bid(proper_skill, country)
        print("Bid for apply: ", my_bid)
        find_element(driver, By.ID, 'BidContent').send_keys(my_bid)
        sleep(2)
        find_element(driver, By.CSS_SELECTOR, 'div[class="wk-submit-block"]').find_element(By.TAG_NAME, 'input').click()
        print("{}th bid was successful!".format(i))
        status = True
        sleep(3)
        break
    
    if status == False:
        print("No match projects.")
        print("Please wait for 2 minutes...")
        print("{} / {} progressed.".format(i - 1, bid_num))
        sleep(120)
        print("Restarting...")
        contract_type += 1
        contract_type %= 2
    else:
        i += 1
    sleep(3)

print("Quiting...")
driver.quit()